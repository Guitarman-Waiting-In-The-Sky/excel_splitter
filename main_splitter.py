from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Alignment

'''
The current working version of this will split one Excel report into many based on unique column values.  Once complete, this will
add the option to split the file as separate tabs within the same report.

'''

def remove_illegal_characters(path):
        
        path=str(path)
        illegal_characters='\/<>:"|?*[]'
        for chars in path:
            if chars in illegal_characters:
                path=path.replace(chars, '_')
        return path


def remove_formula_like_characters(string):
        
        # Currently this doesnt do formulas, this remove the = sign from the beginning so as not to cause errors when opening new file.

        if string[0]=='=':
            string=string[1:]
        
        return string 

def split_into_new_tabs_of_single_report(input_report_path, column_number_to_split_by, sheet=''):

    if sheet=='':
        df = pd.read_excel(input_report_path, sheet_name=0)
    else:
        df = pd.read_excel(input_report_path, sheet_name=str(sheet))

    df = df.fillna('NO_COLUMN_VALUE') # replacing Nan values with NO_COLUMN_VALUE.

    selected_col_values=df.iloc[:,int(column_number_to_split_by)-1].unique()

    final_destination_wb = Workbook()
            
    for unique_items in selected_col_values:

            if unique_items not in df.columns:

                sheet_name=remove_illegal_characters(str(unique_items)) # removing forbidden characters from worksheet names
                sheet_name=sheet_name[:30] # Excel has issues with sheet names over 31 characters, this may be problematic if you have 10 different sheet names with the first 30 chars the same
                final_destination_ws = final_destination_wb.create_sheet(sheet_name)

                row_counter=2
                    
                vectorized_df=df.loc[(df.iloc[:,int(column_number_to_split_by)-1]==unique_items)]

                for i in vectorized_df.itertuples():

                    header_counter=1

                    vectorized_col=vectorized_df.columns

                    for headers in vectorized_df.columns:

                        final_destination_ws.cell(row=1, column=header_counter).value=str(headers) 
                        final_destination_ws.cell(row=1, column=header_counter).alignment =Alignment(horizontal='center', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
                        header_counter=header_counter+1

                    col_counter=1

                    for cols in range(len(vectorized_col)):
                    
                        data=remove_formula_like_characters(str(i[cols+1]))
                        final_destination_ws.cell(row=row_counter, column=col_counter).value=data
                        final_destination_ws.cell(row=row_counter, column=col_counter).alignment =Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
                        col_counter=col_counter+1

                    row_counter=row_counter+1

    del final_destination_wb['Sheet'] # By default xlsx create sheet comes with a tab named sheet.  This gets rid of it.
    final_destination_wb.save('Split_Tab_Report.xlsx')
    final_destination_wb.close()

def split_into_separate_reports(input_report_path, column_number_to_split_by, sheet=''):

    # While pandas unique() is case-sensitive by default, Windows file-saving is not. For now, this script is case-sensitive. If you need two reports for "NAMES" and "NaMeS", Windows wouldn't recognize the
    # case sensitivity in the filename and would just save a single "Names" report.  To get around this, I've created the unique_filename_index to be added to the filename.
    unique_filename_index=2
    previously_used_filenames=[]

    if sheet=='':
        df = pd.read_excel(input_report_path, sheet_name=0) # Sheet 0 = first sheet
    else:
        df = pd.read_excel(input_report_path, sheet_name=str(sheet))

    df = df.fillna('NO_COLUMN_VALUE') # replacing Nan values with NO_COLUMN_VALUE.

    selected_col_values=df.iloc[:,int(column_number_to_split_by)-1].unique()

    for unique_items in selected_col_values:
            
            final_destination_wb = Workbook()
            final_destination_ws = final_destination_wb['Sheet']

            row_counter=2
                
            vectorized_df=df.loc[(df.iloc[:,int(column_number_to_split_by)-1]==unique_items)]

            for i in vectorized_df.itertuples():

                header_counter=1

                vectorized_col=vectorized_df.columns

                for headers in vectorized_df.columns:

                    final_destination_ws.cell(row=1, column=header_counter).value=str(headers)
                    final_destination_ws.cell(row=1, column=header_counter).alignment =Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
                    header_counter=header_counter+1

                col_counter=1

                for cols in range(len(vectorized_col)):
                
                    data=remove_formula_like_characters(str(i[cols+1]))
                    final_destination_ws.cell(row=row_counter, column=col_counter).value=data
                    final_destination_ws.cell(row=row_counter, column=col_counter).alignment =Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
                    col_counter=col_counter+1

                row_counter=row_counter+1


            unique_items=remove_illegal_characters(str(unique_items))

            # Accounting for strings that differ only in case-sensitivity (i.e. "Name" vs. "NAMe")
            # For now, the script is case sensitive

            if unique_items.lower() not in previously_used_filenames:
                previously_used_filenames.append(unique_items.lower())
                final_destination_wb.save(f'{str(unique_items)}.xlsx')
            elif unique_items.lower() in previously_used_filenames:
                final_destination_wb.save(f'{str(unique_items)}_{str(unique_filename_index)}.xlsx')
                unique_filename_index=unique_filename_index+1
            final_destination_wb.close()

